import openpyxl

path = "2010crimedata.xlsx"
workbook = openpyxl.load_workbook(path)

sheet = workbook.active
row = sheet.max_row
col = sheet.max_column

# print("Total Rows:", row)
# print("Total Cols:", col)

switch = {
    2: "City",
    3: "Population",
    4: "Violent Crime",
    5: "Murder & Nonnegligent Manslaughter",
    6: "Rape",
    7: "Robbery",
    8: "Aggravated Assault",
    9: "Property Crime",
    10: "Burglary",
    11: "Larceny Theft",
    12: "Motor Vehicle Theft",
    13: "Arson"
}

lines = []

for y in range(4, 9314 + 1):
    string = "\n\n"
    for x in range(1, col + 1):
        cell_data = sheet.cell(row = y, column = x)
        if(cell_data == 1 and cell_data.value != "None"):
            state = cell_data.value
            string += f"State: {cell_data.value}\n"
        else:
            string += f"{switch.get(x)}: {cell_data.value}\n"
    lines.append(string)

file = open('output_data.txt', 'w')
file.writelines(lines)
file.close()